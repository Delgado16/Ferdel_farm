import csv
import io
import json
from datetime import datetime
from flask import Response, make_response

def exportar_csv(datos, nombre_archivo):
    """Exportar datos a CSV"""
    if not datos:
        return "No hay datos para exportar", 400
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Escribir encabezados
    writer.writerow(datos[0].keys())
    
    # Escribir datos
    for row in datos:
        writer.writerow(row.values())
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response.headers['Content-type'] = 'text/csv'
    return response

def exportar_json(datos, nombre_archivo):
    """Exportar datos a JSON"""
    return Response(
        json.dumps(datos, default=str, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'}
    )

def exportar_excel(datos, nombre_archivo):
    """Exportar datos a Excel (.xlsx) de manera estructurada"""
    if not datos:
        return "No hay datos para exportar", 400
        
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Encabezados
    headers = list(datos[0].keys())
    ws.append(headers)
    
    # Estilo de cabeceras
    header_fill = PatternFill(start_color="2C5E2E", end_color="2C5E2E", fill_type="solid") # primary-color de Ferdel
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Filas de datos
    for row_data in datos:
        row_values = [str(val) if val is not None else "" for val in row_data.values()]
        ws.append(row_values)
        
    # Auto-ajustar el ancho de las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

def exportar_pdf(datos, nombre_archivo):
    """Exportar datos a PDF de manera estructurada y limpia"""
    if not datos:
        return "No hay datos para exportar", 400
        
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    # Usar landscape (apaisado) si hay muchas columnas (> 6) para que quepa bien
    headers = list(datos[0].keys())
    use_landscape = len(headers) > 6
    pagesize = landscape(letter) if use_landscape else letter
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=pagesize,
        rightMargin=36, 
        leftMargin=36, 
        topMargin=36, 
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Título del reporte
    titulo_style = ParagraphStyle(
        'TituloReporte',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#2c5e2e'),
        spaceAfter=15
    )
    
    fecha_style = ParagraphStyle(
        'FechaReporte',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20
    )
    
    # Formatear el nombre del archivo para el título
    titulo_limpio = nombre_archivo.replace('_', ' ').replace('-', ' ').title()
    story.append(Paragraph(f"Reporte de {titulo_limpio}", titulo_style))
    story.append(Paragraph(f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", fecha_style))
    
    # Preparar tabla
    table_data = [headers]
    
    body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    for row in datos:
        row_cells = []
        for val in row.values():
            val_str = str(val) if val is not None else ""
            row_cells.append(Paragraph(val_str, body_style))
        table_data.append(row_cells)
    
    # Ajustar ancho de las columnas proporcionalmente
    col_width = (doc.width) / len(headers)
    t = Table(table_data, colWidths=[col_width] * len(headers))
    
    # Estilo de tabla
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c5e2e')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ])
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response.headers['Content-type'] = 'application/pdf'
    return response
