"""
Módulo de Caja Diaria - Gestión de movimientos y flujo de dinero
"""
from flask import render_template, redirect, url_for, request, flash
from flask_login import current_user
from datetime import datetime, date
from config.database import get_db_cursor
from auth.decorators import admin_required
from . import admin_bp
import io
from flask import make_response
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from helpers.formatters import format_hora

@admin_bp.route('/admin/caja')
@admin_required
def admin_caja():
    """Vista principal de caja - Muestra estado actual"""
    fecha_actual = datetime.now().date()
    
    with get_db_cursor(True) as cursor:
        # Estado de caja (Abierta/Cerrada)
        cursor.execute("""
            SELECT CASE 
                WHEN EXISTS (
                    SELECT 1 FROM caja_movimientos 
                    WHERE Tipo_Movimiento = 'ENTRADA' 
                    AND Descripcion LIKE '%%Apertura%%'
                    AND DATE(Fecha) = %s
                    AND Estado = 'ACTIVO'
                ) THEN 'ABIERTA'
                ELSE 'CERRADA'
            END as estado
        """, (fecha_actual,))
        estado = cursor.fetchone()['estado']
        
        # Resumen del día (solo movimientos ACTIVOS)
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto ELSE 0 END), 0) as entradas,
                COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'SALIDA' THEN Monto ELSE 0 END), 0) as salidas,
                COALESCE(SUM(CASE 
                    WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto 
                    ELSE -Monto 
                END), 0) as saldo_dia
            FROM caja_movimientos
            WHERE DATE(Fecha) = %s
            AND Estado = 'ACTIVO'
        """, (fecha_actual,))
        
        resumen = cursor.fetchone()
        
        # Obtener ventas de ruta en efectivo para agregarlas virtualmente a la caja chica (ya que son efectivo)
        cursor.execute("""
            SELECT COALESCE(SUM(dfr.Total), 0) as total_ruta_efectivo
            FROM facturacion_ruta fr
            INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
            WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
        """, (fecha_actual,))
        total_ruta_efectivo = float(cursor.fetchone()['total_ruta_efectivo'] or 0)
        
        # Movimientos del día
        cursor.execute("""
            SELECT 
                ID_Movimiento,
                Fecha,
                Tipo_Movimiento,
                Descripcion,
                Monto,
                Referencia_Documento,
                Estado
            FROM caja_movimientos
            WHERE DATE(Fecha) = %s
            AND Estado = 'ACTIVO'
            AND (Descripcion NOT LIKE '%%Anulación%%' 
                 AND Descripcion NOT LIKE '%%Contramovimiento%%'
                 AND (Referencia_Documento IS NULL 
                      OR Referencia_Documento NOT LIKE '%%ANUL%%'))
            ORDER BY Fecha DESC
        """, (fecha_actual,))
        
        movimientos = cursor.fetchall()
        
        # Obtener detalle de ventas de ruta para listarlos
        cursor.execute("""
            SELECT 
                fr.ID_FacturaRuta,
                fr.Fecha_Creacion AS Fecha,
                u.NombreUsuario AS Vendedor,
                SUM(COALESCE(dfr.Total, 0)) AS Monto
            FROM facturacion_ruta fr
            INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
            INNER JOIN asignacion_vendedores av ON fr.ID_Asignacion = av.ID_Asignacion
            INNER JOIN usuarios u ON av.ID_Usuario = u.ID_Usuario
            WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
            GROUP BY fr.ID_FacturaRuta, fr.Fecha_Creacion, u.NombreUsuario
            ORDER BY fr.Fecha_Creacion DESC
        """, (fecha_actual,))
        movimientos_ruta = cursor.fetchall()
        
        lista_movimientos = list(movimientos)
        for mr in movimientos_ruta:
            lista_movimientos.append({
                'ID_Movimiento': f"RUT-{mr['ID_FacturaRuta']}",
                'Fecha': mr['Fecha'],
                'Tipo_Movimiento': 'ENTRADA',
                'Descripcion': f"Venta de Ruta al contado - Vendedor: {mr['Vendedor']}",
                'Monto': float(mr['Monto']),
                'Referencia_Documento': f"RUT-{mr['ID_FacturaRuta']:05d}",
                'Estado': 'ACTIVO'
            })
        
        lista_movimientos.sort(key=lambda x: x['Fecha'], reverse=True)
    
    entradas_totales = float(resumen['entradas'] or 0) + total_ruta_efectivo
    saldo_dia_total = float(resumen['saldo_dia'] or 0) + total_ruta_efectivo
    
    datos = {
        'fecha': fecha_actual.strftime('%d/%m/%Y'),
        'estado': estado,
        'entradas': entradas_totales,
        'salidas': float(resumen['salidas'] or 0),
        'saldo_dia': saldo_dia_total,
        'movimientos': lista_movimientos
    }
    
    return render_template('admin/caja/caja.html', caja=datos)

@admin_bp.route('/admin/caja/aperturar', methods=['POST'])
@admin_required
def admin_caja_aperturar():
    """Abre la caja con un monto inicial"""
    try:
        monto = float(request.form.get('monto_inicial', 0))
        
        if monto < 0:
            flash('El monto debe ser mayor o igual a 0', 'error')
            return redirect(url_for('admin.admin_caja'))
        
        fecha_actual = datetime.now().date()
        
        with get_db_cursor(True) as cursor:
            # Verificar si ya hay apertura hoy
            cursor.execute("""
                SELECT 1 FROM caja_movimientos 
                WHERE Tipo_Movimiento = 'ENTRADA' 
                AND Descripcion LIKE '%%Apertura%%'
                AND DATE(Fecha) = %s
                AND Estado = 'ACTIVO'
                LIMIT 1
            """, (fecha_actual,))
            
            if cursor.fetchone():
                flash('La caja ya está aperturada hoy', 'error')
                return redirect(url_for('admin.admin_caja'))
            
            # Registrar apertura
            cursor.execute("""
                INSERT INTO caja_movimientos 
                (Fecha, Tipo_Movimiento, Descripcion, Monto, ID_Usuario, Estado)
                VALUES (NOW(), 'ENTRADA', %s, %s, %s, 'ACTIVO')
            """, (f"Apertura de caja", monto, current_user.id))
            
            flash(f'Caja aperturada con C${monto:.2f}', 'success')
            return redirect(url_for('admin.admin_caja'))
            
    except ValueError:
        flash('Monto inválido', 'error')
        return redirect(url_for('admin.admin_caja'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))

@admin_bp.route('/admin/caja/movimiento', methods=['POST'])
@admin_required
def admin_caja_movimiento():
    """Registra un movimiento manual de entrada o salida"""
    try:
        tipo = request.form.get('tipo_movimiento')
        descripcion = request.form.get('descripcion', '').strip()
        monto = float(request.form.get('monto', 0))
        referencia = request.form.get('referencia_documento', '').strip()
        
        # Validaciones básicas
        if tipo not in ['ENTRADA', 'SALIDA']:
            flash('Tipo de movimiento inválido', 'error')
            return redirect(url_for('admin.admin_caja'))
        
        if monto <= 0:
            flash('El monto debe ser mayor a 0', 'error')
            return redirect(url_for('admin.admin_caja'))
        
        if not descripcion:
            flash('Descripción requerida', 'error')
            return redirect(url_for('admin.admin_caja'))
        
        with get_db_cursor(True) as cursor:
            # Para salidas, verificar que la caja esté abierta
            if tipo == 'SALIDA':
                fecha_actual = datetime.now().date()
                cursor.execute("""
                    SELECT 1 FROM caja_movimientos 
                    WHERE Tipo_Movimiento = 'ENTRADA' 
                    AND Descripcion LIKE '%%Apertura%%'
                    AND DATE(Fecha) = %s
                    AND Estado = 'ACTIVO'
                    LIMIT 1
                """, (fecha_actual,))
                
                if not cursor.fetchone():
                    flash('La caja no está aperturada', 'error')
                    return redirect(url_for('admin.admin_caja'))
            
            # Registrar movimiento
            cursor.execute("""
                INSERT INTO caja_movimientos 
                (Fecha, Tipo_Movimiento, Descripcion, Monto, Referencia_Documento, ID_Usuario, Estado)
                VALUES (NOW(), %s, %s, %s, %s, %s, 'ACTIVO')
            """, (tipo, descripcion, monto, referencia, current_user.id))
            
            flash(f'Movimiento registrado: {descripcion}', 'success')
            return redirect(url_for('admin.admin_caja'))
            
    except ValueError:
        flash('Monto inválido', 'error')
        return redirect(url_for('admin.admin_caja'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))

@admin_bp.route('/admin/caja/cerrar', methods=['POST'])
@admin_required
def admin_caja_cerrar():
    """Cierra la caja del día"""
    try:
        with get_db_cursor(True) as cursor:
            fecha_actual = datetime.now().date()
            
            # Obtener saldo de cierre
            cursor.execute("""
                SELECT COALESCE(SUM(CASE 
                    WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto 
                    ELSE -Monto 
                END), 0) as saldo_final
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s AND Estado = 'ACTIVO'
            """, (fecha_actual,))
            
            saldo = cursor.fetchone()['saldo_final'] or 0
            
            # Registrar cierre
            cursor.execute("""
                INSERT INTO caja_movimientos 
                (Fecha, Tipo_Movimiento, Descripcion, Monto, ID_Usuario, Estado)
                VALUES (NOW(), 'SALIDA', %s, %s, %s, 'ACTIVO')
            """, ("Cierre de caja", saldo, current_user.id))
            
            flash(f'Caja cerrada. Saldo final: C${saldo:.2f}', 'success')
            return redirect(url_for('admin.admin_caja'))
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))

@admin_bp.route('/admin/caja/anular/<int:id_movimiento>', methods=['POST'])
@admin_required
def admin_caja_anular(id_movimiento):
    """Anula un movimiento de caja"""
    try:
        with get_db_cursor(True) as cursor:
            # Obtener movimiento
            cursor.execute("""
                SELECT Monto, Tipo_Movimiento FROM caja_movimientos
                WHERE ID_Movimiento = %s AND Estado = 'ACTIVO'
            """, (id_movimiento,))
            
            movimiento = cursor.fetchone()
            if not movimiento:
                flash('Movimiento no encontrado', 'error')
                return redirect(url_for('admin.admin_caja'))
            
            # Marcar como anulado
            cursor.execute("""
                UPDATE caja_movimientos
                SET Estado = 'ANULADO'
                WHERE ID_Movimiento = %s
            """, (id_movimiento,))
            
            # Registrar contramovimiento
            tipo_inverso = 'SALIDA' if movimiento['Tipo_Movimiento'] == 'ENTRADA' else 'ENTRADA'
            cursor.execute("""
                INSERT INTO caja_movimientos 
                (Fecha, Tipo_Movimiento, Descripcion, Monto, Referencia_Documento, ID_Usuario, Estado)
                VALUES (NOW(), %s, %s, %s, %s, %s, 'ACTIVO')
            """, (tipo_inverso, f"Anulación de movimiento {id_movimiento}", 
                  movimiento['Monto'], f"ANUL-{id_movimiento}", current_user.id))
            
            flash('Movimiento anulado', 'success')
            return redirect(url_for('admin.admin_caja'))
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))

@admin_bp.route('/admin/caja/historial')
@admin_required
def admin_caja_historial():
    """Muestra historial de movimientos de caja con filtro por fecha (hoy por defecto)"""
    try:
        from datetime import datetime
        
        # Obtener fecha del parámetro GET (si no viene, usar hoy)
        fecha_param = request.args.get('fecha', '')
        
        with get_db_cursor() as cursor:
            # Si no se especificó fecha, usar la fecha actual
            if not fecha_param:
                cursor.execute("SELECT CURDATE() as hoy")
                fecha_actual = cursor.fetchone()
                fecha_param = fecha_actual['hoy'].strftime('%Y-%m-%d')
            
            # Validar que la fecha no sea futura
            fecha_maxima = datetime.now().strftime('%Y-%m-%d')
            if fecha_param > fecha_maxima:
                fecha_param = fecha_maxima
            
            # Obtener movimientos de la fecha seleccionada
            cursor.execute("""
                SELECT 
                    ID_Movimiento,
                    Fecha,
                    Tipo_Movimiento,
                    Descripcion,
                    Monto,
                    Estado,
                    Referencia_Documento,
                    ID_Factura,
                    ID_Pagos_cxc
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s
                ORDER BY Fecha ASC
            """, (fecha_param,))
            
            movimientos = cursor.fetchall()
            
            # Calcular totales de la fecha seleccionada
            entradas = sum(m['Monto'] for m in movimientos 
                          if m['Tipo_Movimiento'] == 'ENTRADA' and m['Estado'] != 'ANULADO')
            salidas = sum(m['Monto'] for m in movimientos 
                         if m['Tipo_Movimiento'] == 'SALIDA' and m['Estado'] != 'ANULADO')
            saldo_dia = entradas - salidas
            
            # Contar movimientos
            total = len([m for m in movimientos if m['Estado'] != 'ANULADO'])
            total_anulados = len([m for m in movimientos if m['Estado'] == 'ANULADO'])
            
            # Obtener fechas disponibles para el selector (últimos 30 días con movimientos)
            cursor.execute("""
                SELECT DISTINCT DATE(Fecha) as fecha
                FROM caja_movimientos
                ORDER BY fecha DESC
                LIMIT 30
            """)
            fechas_disponibles = cursor.fetchall()
            
            # Determinar estado de la caja basado en movimientos (sin tabla caja_estados)
            # Si hay movimientos de apertura/cierre en la fecha
            tiene_apertura = any(m.get('Referencia_Documento') == 'APERTURA' for m in movimientos)
            tiene_cierre = any(m.get('Referencia_Documento') == 'CIERRE' for m in movimientos)
            
            if tiene_apertura and not tiene_cierre:
                estado = 'ABIERTA'
            elif tiene_apertura and tiene_cierre:
                estado = 'CERRADA'
            else:
                estado = 'SIN APERTURA'
            
            # Formatear fecha para mostrar
            fecha_obj = datetime.strptime(fecha_param, '%Y-%m-%d')
            fecha_formateada = fecha_obj.strftime('%d/%m/%Y')
        
        return render_template('admin/caja/historial.html',
                             movimientos=movimientos,
                             fecha=fecha_param,
                             fecha_formateada=fecha_formateada,
                             fecha_maxima=fecha_maxima,
                             fechas_disponibles=fechas_disponibles,
                             entradas=entradas,
                             salidas=salidas,
                             saldo_dia=saldo_dia,
                             total=total,
                             total_anulados=total_anulados,
                             estado=estado)
        
    except Exception as e:
        flash(f'Error al cargar historial: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))

@admin_bp.route('/admin/caja/reporte')
@admin_required
def admin_caja_reporte():
    """Reporte de caja por rango de fechas"""
    try:
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        
        if fecha_inicio_str and fecha_fin_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            
            # Asegurar que fecha_inicio sea menor o igual a fecha_fin
            if fecha_inicio > fecha_fin:
                fecha_inicio, fecha_fin = fecha_fin, fecha_inicio
            
            with get_db_cursor(True) as cursor:
                # Reporte agrupado por día
                cursor.execute("""
                    SELECT 
                        DATE(Fecha) as fecha,
                        COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto ELSE 0 END), 0) as entradas,
                        COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'SALIDA' THEN Monto ELSE 0 END), 0) as salidas,
                        COUNT(*) as movimientos
                    FROM caja_movimientos
                    WHERE DATE(Fecha) BETWEEN %s AND %s
                    AND Estado = 'ACTIVO'
                    GROUP BY DATE(Fecha)
                    ORDER BY fecha DESC
                """, (fecha_inicio, fecha_fin))
                
                reporte = cursor.fetchall()
                
                # Totales generales del período
                cursor.execute("""
                    SELECT 
                        COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto ELSE 0 END), 0) as entradas_total,
                        COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'SALIDA' THEN Monto ELSE 0 END), 0) as salidas_total,
                        COUNT(*) as total_movimientos
                    FROM caja_movimientos
                    WHERE DATE(Fecha) BETWEEN %s AND %s
                    AND Estado = 'ACTIVO'
                """, (fecha_inicio, fecha_fin))
                
                totales = cursor.fetchone()
            
            return render_template('admin/caja/reporte.html',
                                 fecha_inicio=fecha_inicio.strftime('%Y-%m-%d'),
                                 fecha_fin=fecha_fin.strftime('%Y-%m-%d'),
                                 reporte=reporte,
                                 entradas_total=float(totales['entradas_total'] or 0),
                                 salidas_total=float(totales['salidas_total'] or 0),
                                 movimientos_total=totales['total_movimientos'] or 0)
        
        return render_template('admin/caja/reporte.html')
            
    except ValueError:
        flash('Fechas inválidas', 'error')
        return redirect(url_for('admin.admin_caja_reporte'))
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin.admin_caja'))


@admin_bp.route('/admin/caja/exportar/excel')
@admin_required
def admin_caja_exportar_excel():
    try:
        fecha_str = request.args.get('fecha')
        if fecha_str:
            fecha_actual = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha_actual = datetime.now().date()
            
        with get_db_cursor(True) as cursor:
            # Estado de caja
            cursor.execute("""
                SELECT CASE 
                    WHEN EXISTS (
                        SELECT 1 FROM caja_movimientos 
                        WHERE Tipo_Movimiento = 'ENTRADA' 
                        AND Descripcion LIKE '%%Apertura%%'
                        AND DATE(Fecha) = %s
                        AND Estado = 'ACTIVO'
                    ) THEN 'ABIERTA'
                    ELSE 'CERRADA'
                END as estado
            """, (fecha_actual,))
            estado = cursor.fetchone()['estado']
            
            # Resumen del día
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto ELSE 0 END), 0) as entradas,
                    COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'SALIDA' THEN Monto ELSE 0 END), 0) as salidas,
                    COALESCE(SUM(CASE 
                        WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto 
                        ELSE -Monto 
                    END), 0) as saldo_dia
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s
                AND Estado = 'ACTIVO'
            """, (fecha_actual,))
            resumen = cursor.fetchone()
            
            # Obtener ventas de ruta en efectivo para agregarlas virtualmente a la caja chica (ya que son efectivo)
            cursor.execute("""
                SELECT COALESCE(SUM(dfr.Total), 0) as total_ruta_efectivo
                FROM facturacion_ruta fr
                INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
                WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
            """, (fecha_actual,))
            total_ruta_efectivo = float(cursor.fetchone()['total_ruta_efectivo'] or 0)
            
            # Movimientos
            cursor.execute("""
                SELECT 
                    Fecha,
                    Tipo_Movimiento,
                    Descripcion,
                    Monto,
                    Referencia_Documento
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s
                AND Estado = 'ACTIVO'
                AND (Descripcion NOT LIKE '%%Anulación%%' 
                     AND Descripcion NOT LIKE '%%Contramovimiento%%'
                     AND (Referencia_Documento IS NULL 
                          OR Referencia_Documento NOT LIKE '%%ANUL%%'))
                ORDER BY Fecha ASC
            """, (fecha_actual,))
            movimientos = cursor.fetchall()
            
            # Obtener detalle de ventas de ruta para listarlos
            cursor.execute("""
                SELECT 
                    fr.ID_FacturaRuta,
                    fr.Fecha_Creacion AS Fecha,
                    u.NombreUsuario AS Vendedor,
                    SUM(COALESCE(dfr.Total, 0)) AS Monto
                FROM facturacion_ruta fr
                INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
                INNER JOIN asignacion_vendedores av ON fr.ID_Asignacion = av.ID_Asignacion
                INNER JOIN usuarios u ON av.ID_Usuario = u.ID_Usuario
                WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
                GROUP BY fr.ID_FacturaRuta, fr.Fecha_Creacion, u.NombreUsuario
                ORDER BY fr.Fecha_Creacion ASC
            """, (fecha_actual,))
            movimientos_ruta = cursor.fetchall()
            
            lista_movimientos = []
            for m in movimientos:
                lista_movimientos.append({
                    'Fecha': m['Fecha'],
                    'Tipo_Movimiento': m['Tipo_Movimiento'],
                    'Descripcion': m['Descripcion'],
                    'Monto': float(m['Monto']),
                    'Referencia_Documento': m['Referencia_Documento']
                })
                
            for mr in movimientos_ruta:
                lista_movimientos.append({
                    'Fecha': mr['Fecha'],
                    'Tipo_Movimiento': 'ENTRADA',
                    'Descripcion': f"Venta de Ruta al contado - Vendedor: {mr['Vendedor']}",
                    'Monto': float(mr['Monto']),
                    'Referencia_Documento': f"RUT-{mr['ID_FacturaRuta']:05d}"
                })
                
            lista_movimientos.sort(key=lambda x: x['Fecha'])
            
            resumen_entradas = float(resumen['entradas'] or 0) + total_ruta_efectivo
            resumen_saldo_dia = float(resumen['saldo_dia'] or 0) + total_ruta_efectivo
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flujo de Caja"
        
        ws.views.sheetView[0].showGridLines = True
        
        font_family = "Segoe UI"
        
        fill_header = PatternFill(start_color="2C5E2E", end_color="2C5E2E", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_entrada = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        fill_salida = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        fill_total = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        
        font_title = Font(name=font_family, size=16, bold=True, color="2C5E2E")
        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_bold = Font(name=font_family, size=11, bold=True)
        font_regular = Font(name=font_family, size=11)
        font_small = Font(name=font_family, size=9, italic=True, color="555555")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws["A1"] = "REPORTE DIARIO DE FLUJO DE CAJA"
        ws["A1"].font = font_title
        ws["A2"] = f"Fecha: {fecha_actual.strftime('%d/%m/%Y')} | Estado de Caja: {estado}"
        ws["A2"].font = font_small
        
        ws["A4"] = "RESUMEN"
        ws["A4"].font = font_bold
        ws.merge_cells("A4:B4")
        
        ws["A5"] = "Total Entradas"
        ws["B5"] = resumen_entradas
        ws["B5"].number_format = '"C$"#,##0.00'
        ws["A5"].fill = fill_entrada
        ws["B5"].fill = fill_entrada
        ws["A5"].font = font_regular
        ws["B5"].font = font_bold
        
        ws["A6"] = "Total Salidas"
        ws["B6"] = float(resumen['salidas'] or 0)
        ws["B6"].number_format = '"C$"#,##0.00'
        ws["A6"].fill = fill_salida
        ws["B6"].fill = fill_salida
        ws["A6"].font = font_regular
        ws["B6"].font = font_bold
        
        ws["A7"] = "Saldo Neto"
        ws["B7"] = resumen_saldo_dia
        ws["B7"].number_format = '"C$"#,##0.00'
        ws["A7"].fill = fill_total
        ws["B7"].fill = fill_total
        ws["A7"].font = font_bold
        ws["B7"].font = font_bold
        
        for r in range(5, 8):
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=2).border = thin_border
            
        headers = ["Hora", "Tipo", "Descripción", "Referencia", "Monto"]
        start_row = 10
        
        ws.cell(row=start_row - 1, column=1, value="DETALLE DE MOVIMIENTOS").font = font_bold
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        from helpers.formatters import format_hora
        
        current_row = start_row + 1
        for idx, mov in enumerate(lista_movimientos):
            hora_str = format_hora(mov['Fecha'])
            c_hora = ws.cell(row=current_row, column=1, value=hora_str)
            c_hora.alignment = Alignment(horizontal="center")
            
            c_tipo = ws.cell(row=current_row, column=2, value=mov['Tipo_Movimiento'])
            c_tipo.alignment = Alignment(horizontal="center")
            if mov['Tipo_Movimiento'] == 'ENTRADA':
                c_tipo.font = Font(name=font_family, size=11, bold=True, color="1B5E20")
            else:
                c_tipo.font = Font(name=font_family, size=11, bold=True, color="B71C1C")
                
            c_desc = ws.cell(row=current_row, column=3, value=mov['Descripcion'])
            
            c_ref = ws.cell(row=current_row, column=4, value=mov['Referencia_Documento'] or "")
            c_ref.alignment = Alignment(horizontal="center")
            
            c_monto = ws.cell(row=current_row, column=5, value=float(mov['Monto'] or 0))
            c_monto.number_format = '"C$"#,##0.00'
            c_monto.alignment = Alignment(horizontal="right")
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.font = font_regular if col_idx != 2 else cell.font
                if idx % 2 == 1:
                    cell.fill = fill_zebra
                    
            current_row += 1
            
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        
        import io
        from flask import make_response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=flujo_caja_{fecha_actual.strftime("%Y%m%d")}.xlsx'
        response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    except Exception as e:
        flash(f"Error al exportar Excel: {str(e)}", "error")
        return redirect(url_for('admin.admin_caja'))


@admin_bp.route('/admin/caja/exportar/pdf')
@admin_required
def admin_caja_exportar_pdf():
    try:
        fecha_str = request.args.get('fecha')
        if fecha_str:
            fecha_actual = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha_actual = datetime.now().date()
            
        with get_db_cursor(True) as cursor:
            # Estado de caja
            cursor.execute("""
                SELECT CASE 
                    WHEN EXISTS (
                        SELECT 1 FROM caja_movimientos 
                        WHERE Tipo_Movimiento = 'ENTRADA' 
                        AND Descripcion LIKE '%%Apertura%%'
                        AND DATE(Fecha) = %s
                        AND Estado = 'ACTIVO'
                    ) THEN 'ABIERTA'
                    ELSE 'CERRADA'
                END as estado
            """, (fecha_actual,))
            estado = cursor.fetchone()['estado']
            
            # Resumen del día
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto ELSE 0 END), 0) as entradas,
                    COALESCE(SUM(CASE WHEN Tipo_Movimiento = 'SALIDA' THEN Monto ELSE 0 END), 0) as salidas,
                    COALESCE(SUM(CASE 
                        WHEN Tipo_Movimiento = 'ENTRADA' THEN Monto 
                        ELSE -Monto 
                    END), 0) as saldo_dia
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s
                AND Estado = 'ACTIVO'
            """, (fecha_actual,))
            resumen = cursor.fetchone()
            
            # Obtener ventas de ruta en efectivo para agregarlas virtualmente a la caja chica (ya que son efectivo)
            cursor.execute("""
                SELECT COALESCE(SUM(dfr.Total), 0) as total_ruta_efectivo
                FROM facturacion_ruta fr
                INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
                WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
            """, (fecha_actual,))
            total_ruta_efectivo = float(cursor.fetchone()['total_ruta_efectivo'] or 0)
            
            # Movimientos
            cursor.execute("""
                SELECT 
                    Fecha,
                    Tipo_Movimiento,
                    Descripcion,
                    Monto,
                    Referencia_Documento
                FROM caja_movimientos
                WHERE DATE(Fecha) = %s
                AND Estado = 'ACTIVO'
                AND (Descripcion NOT LIKE '%%Anulación%%' 
                     AND Descripcion NOT LIKE '%%Contramovimiento%%'
                     AND (Referencia_Documento IS NULL 
                          OR Referencia_Documento NOT LIKE '%%ANUL%%'))
                ORDER BY ID_Movimiento ASC
            """, (fecha_actual,))
            movimientos = cursor.fetchall()
            
            # Obtener detalle de ventas de ruta para listarlos
            cursor.execute("""
                SELECT 
                    fr.ID_FacturaRuta,
                    fr.Fecha_Creacion AS Fecha,
                    u.NombreUsuario AS Vendedor,
                    SUM(COALESCE(dfr.Total, 0)) AS Monto
                FROM facturacion_ruta fr
                INNER JOIN detalle_facturacion_ruta dfr ON fr.ID_FacturaRuta = dfr.ID_FacturaRuta
                INNER JOIN asignacion_vendedores av ON fr.ID_Asignacion = av.ID_Asignacion
                INNER JOIN usuarios u ON av.ID_Usuario = u.ID_Usuario
                WHERE DATE(fr.Fecha_Creacion) = %s AND fr.Estado = 'Activa' AND fr.Credito_Contado = 1
                GROUP BY fr.ID_FacturaRuta, fr.Fecha_Creacion, u.NombreUsuario
                ORDER BY fr.Fecha_Creacion ASC
            """, (fecha_actual,))
            movimientos_ruta = cursor.fetchall()
            
            lista_movimientos = []
            for m in movimientos:
                lista_movimientos.append({
                    'Fecha': m['Fecha'],
                    'Tipo_Movimiento': m['Tipo_Movimiento'],
                    'Descripcion': m['Descripcion'],
                    'Monto': float(m['Monto']),
                    'Referencia_Documento': m['Referencia_Documento']
                })
                
            for mr in movimientos_ruta:
                lista_movimientos.append({
                    'Fecha': mr['Fecha'],
                    'Tipo_Movimiento': 'ENTRADA',
                    'Descripcion': f"Venta de Ruta al contado - Vendedor: {mr['Vendedor']}",
                    'Monto': float(mr['Monto']),
                    'Referencia_Documento': f"RUT-{mr['ID_FacturaRuta']:05d}"
                })
                
            lista_movimientos.sort(key=lambda x: x['Fecha'])
            
            resumen_entradas = float(resumen['entradas'] or 0) + total_ruta_efectivo
            resumen_saldo_dia = float(resumen['saldo_dia'] or 0) + total_ruta_efectivo
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            rightMargin=36, 
            leftMargin=36, 
            topMargin=36, 
            bottomMargin=36
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'PdfTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#2c5e2e'),
            spaceAfter=5
        )
        subtitle_style = ParagraphStyle(
            'PdfSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#475569'),
            spaceAfter=15
        )
        section_style = ParagraphStyle(
            'PdfSection',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=colors.HexColor('#1e293b'),
            spaceBefore=12,
            spaceAfter=6
        )
        cell_style = ParagraphStyle(
            'PdfCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            leading=11
        )
        cell_bold_style = ParagraphStyle(
            'PdfCellBold',
            parent=cell_style,
            fontName='Helvetica-Bold'
        )
        cell_right_style = ParagraphStyle(
            'PdfCellRight',
            parent=cell_style,
            alignment=2
        )
        cell_center_style = ParagraphStyle(
            'PdfCellCenter',
            parent=cell_style,
            alignment=1
        )
        
        story.append(Paragraph("REPORTE DIARIO DE FLUJO DE CAJA - FERDEL", title_style))
        story.append(Paragraph(f"Fecha de Reporte: {fecha_actual.strftime('%d/%m/%Y')} | Estado de Caja: {estado}", subtitle_style))
        
        story.append(Paragraph("RESUMEN DE CAJA", section_style))
        resumen_data = [
            [Paragraph("Concepto", cell_bold_style), Paragraph("Monto (C$)", cell_bold_style)],
            [Paragraph("Total Entradas", cell_style), Paragraph(f"C$ {resumen_entradas:,.2f}", cell_style)],
            [Paragraph("Total Salidas", cell_style), Paragraph(f"C$ {float(resumen['salidas'] or 0):,.2f}", cell_style)],
            [Paragraph("Saldo Neto", cell_bold_style), Paragraph(f"C$ {resumen_saldo_dia:,.2f}", cell_bold_style)]
        ]
        
        t_resumen = Table(resumen_data, colWidths=[200, 150])
        t_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor('#e2e8f0')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0,1), (1,1), colors.HexColor('#E6F4EA')),
            ('BACKGROUND', (0,2), (1,2), colors.HexColor('#FCE8E6')),
            ('BACKGROUND', (0,3), (1,3), colors.HexColor('#E8F0FE')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t_resumen)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("DETALLE DE MOVIMIENTOS", section_style))
        
        mov_headers = ["Hora", "Tipo", "Descripción", "Referencia", "Monto"]
        table_data = [[Paragraph(h, cell_bold_style) for h in mov_headers]]
        
        for idx, mov in enumerate(lista_movimientos):
            hora_str = format_hora(mov['Fecha'])
            tipo = mov['Tipo_Movimiento']
            monto_val = f"C$ {float(mov['Monto'] or 0):,.2f}"
            
            row = [
                Paragraph(hora_str, cell_center_style),
                Paragraph(tipo, ParagraphStyle('tipo_st', parent=cell_center_style, textColor=colors.HexColor('#1B5E20') if tipo == 'ENTRADA' else colors.HexColor('#B71C1C'), fontName='Helvetica-Bold')),
                Paragraph(mov['Descripcion'], cell_style),
                Paragraph(mov['Referencia_Documento'] or "", cell_center_style),
                Paragraph(monto_val, cell_right_style)
            ]
            table_data.append(row)
            
        t_mov = Table(table_data, colWidths=[65, 75, 200, 110, 90])
        
        row_styles = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c5e2e')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]
        
        for col_idx in range(len(mov_headers)):
            table_data[0][col_idx].style.textColor = colors.white
            
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                row_styles.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc')))
                
        t_mov.setStyle(TableStyle(row_styles))
        story.append(t_mov)
        
        doc.build(story)
        buffer.seek(0)
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=flujo_caja_{fecha_actual.strftime("%Y%m%d")}.pdf'
        response.headers['Content-type'] = 'application/pdf'
        return response
    except Exception as e:
        flash(f"Error al exportar PDF: {str(e)}", "error")
        return redirect(url_for('admin.admin_caja'))